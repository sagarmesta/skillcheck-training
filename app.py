from flask import Flask, render_template, request, redirect, session, send_file
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
from users import users
from threading import Lock

app = Flask(__name__)
app.secret_key = "secret123"

DATA_FOLDER = "data"
LOG_FILE = "user_records.xlsx"

os.makedirs(DATA_FOLDER, exist_ok=True)

# 🔒 LOCK FOR MULTI-USER SAFETY
lock = Lock()

# ---------- LOGIN LOG FILE ----------
def get_logbook():
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["User", "Login Time", "Logout Time"])
        wb.save(LOG_FILE)
    return load_workbook(LOG_FILE)

# ---------- USER FILE ----------
def create_user_file(user):
    file = f"{DATA_FOLDER}/{user}.xlsx"
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "User", "WorkBook", "Start Date", "End Date", "Form No",
            "Brand", "Manufacturer", "Model", "Commander",
            "Destination", "Reference No", "Part No",
            "Part Description", "Description", "Remarks"
        ])
        with lock:
            wb.save(file)

# ---------- FORM NO ----------
def get_form_no(user):
    file = f"{DATA_FOLDER}/{user}.xlsx"
    if not os.path.exists(file):
        return 0
    wb = load_workbook(file)
    ws = wb.active
    return ws.max_row - 1

# ---------- LOGIN ----------
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form["username"]
        p = request.form["password"]

        if u in users and users[u] == p:
            session["user"] = u
            create_user_file(u)

            wb = get_logbook()
            ws = wb.active
            ws.append([u, datetime.now(), None])

            with lock:
                wb.save(LOG_FILE)

            return redirect("/admin" if u == "admin" else "/dashboard")

        return "Invalid Login ❌"

    return render_template("login.html", users=users)

# ---------- DASHBOARD ----------
@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user" not in session or session["user"] == "admin":
        return redirect("/")

    user = session["user"]
    file = f"{DATA_FOLDER}/{user}.xlsx"

    if request.method == "POST":
        wb = load_workbook(file)
        ws = wb.active

        form_no = ws.max_row - 1
        d = request.form

        ws.append([
            user,
            "TRAINING",
            d.get("start_date"),
            d.get("end_date"),
            form_no,
            d.get("brand"),
            d.get("manufacturer"),
            d.get("model"),
            d.get("commander"),
            d.get("destination"),
            d.get("ref"),
            d.get("part_no"),
            d.get("part_desc"),
            d.get("desc"),
            d.get("remarks")
        ])

        with lock:
            wb.save(file)

        return redirect("/dashboard?success=1")

    return render_template("dashboard.html",
                           form_no=get_form_no(user),
                           success=request.args.get("success"))

# ---------- ADMIN ----------
@app.route("/admin")
def admin():
    if session.get("user") != "admin":
        return redirect("/")

    wb = get_logbook()
    ws = wb.active

    data = []
    active = 0

    logs = list(ws.iter_rows(min_row=2, values_only=True))

    for user in users.keys():
        if user == "admin":
            continue

        form_no = get_form_no(user)
        user_logs = [r for r in logs if r[0] == user]

        if user_logs:
            last = user_logs[-1]
            login_time = last[1]
            logout_time = last[2]

            if logout_time is None:
                active += 1
        else:
            login_time = "-"
            logout_time = "-"

        data.append([user, login_time, logout_time, form_no])

    return render_template("admin.html",
                           data=data,
                           total=len(data),
                           active=active,
                           users=users)

# ---------- DELETE LOGS ----------
@app.route("/delete_logs_user", methods=["POST"])
def delete_logs_user():
    if session.get("user") != "admin":
        return redirect("/")

    username = request.form.get("username")

    wb = get_logbook()
    ws = wb.active

    remaining = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != username:
            remaining.append(row)

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(["User", "Login Time", "Logout Time"])

    for r in remaining:
        new_ws.append(r)

    with lock:
        new_wb.save(LOG_FILE)

    return redirect("/admin")

# ---------- RESET ----------
@app.route("/reset_form_no", methods=["POST"])
def reset_form_no():
    if session.get("user") != "admin":
        return redirect("/")

    for file in os.listdir(DATA_FOLDER):
        path = os.path.join(DATA_FOLDER, file)

        wb = Workbook()
        ws = wb.active

        ws.append([
            "User", "WorkBook", "Start Date", "End Date", "Form No",
            "Brand", "Manufacturer", "Model", "Commander",
            "Destination", "Reference No", "Part No",
            "Part Description", "Description", "Remarks"
        ])

        with lock:
            wb.save(path)

    return redirect("/admin")

# ---------- DOWNLOAD ----------
@app.route("/download/<user>")
def download_user(user):
    if session.get("user") != "admin":
        return redirect("/")
    return send_file(f"{DATA_FOLDER}/{user}.xlsx", as_attachment=True)

@app.route("/download_logs")
def download_logs():
    if session.get("user") != "admin":
        return redirect("/")
    return send_file(LOG_FILE, as_attachment=True)

# ---------- CHANGE PASSWORD ----------
@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    if "user" not in session:
        return redirect("/")

    msg = ""

    if request.method == "POST":
        old = request.form.get("old")
        new = request.form.get("new")
        user = session["user"]

        if users.get(user) == old:
            users[user] = new
            with open("users.py", "w") as f:
                f.write("users = " + str(users))
            msg = "Password Updated Successfully ✅"
        else:
            msg = "Old password incorrect ❌"

    return render_template("change_password.html", msg=msg)

# ---------- LOGOUT ----------
@app.route("/logout")
def logout():
    if "user" in session:
        wb = get_logbook()
        ws = wb.active

        for i in range(ws.max_row, 1, -1):
            if ws.cell(i, 1).value == session["user"] and ws.cell(i, 3).value is None:
                ws.cell(i, 3).value = datetime.now()
                break

        with lock:
            wb.save(LOG_FILE)

    session.clear()
    return redirect("/")

# ---------- RUN ----------
if __name__ == "__main__":
    print("🔥 Server Running...")
    app.run()
